"""
小红书博主博文获取工具
使用Playwright模拟浏览器行为
安装依赖: pip install playwright openpyxl
安装浏览器: playwright install chromium

执行流程：
1. 进入主页，点击第一篇博文触发登录
2. 等待用户手动登录
3. 登录后滚动获取所有帖子信息，存储到Excel
4. 根据Excel记录遍历处理每篇帖子
5. 每篇博文创建一个文件夹，包含图片和markdown文件

配置说明：
- 所有参数请在 config.py 中修改
- 无需修改本文件代码
"""

import json
import time
import re
import os
import base64
from typing import List, Dict, Optional, Set
from dataclasses import dataclass, asdict
from playwright.sync_api import sync_playwright, Page, ElementHandle
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment

# 导入配置
import config


@dataclass
class NoteData:
    """笔记数据结构"""
    index: int
    note_id: str
    title: str
    content: str
    images: List[str]
    videos: List[str]  # 视频地址列表
    url: str
    likes: str = ""
    collects: str = ""
    comments: str = ""
    is_video: bool = False  # 是否为视频帖子


class XiaoHongShuScraper:
    """小红书爬虫类"""
    
    def __init__(self, headless: bool = None, output_dir: str = None, download_images: bool = None, max_notes: int = None, existing_note_ids: Set[str] = None, excel_suffix: str = None):
        # 从配置文件读取参数，参数传入优先于配置文件
        self.headless = headless if headless is not None else config.HEADLESS
        self.output_dir = output_dir if output_dir is not None else config.OUTPUT_DIR
        self.download_images = download_images if download_images is not None else config.DOWNLOAD_IMAGES
        self.max_notes = max_notes
        self.existing_note_ids = existing_note_ids or set()  # 已存在的 note_id 集合（用于增量爬取）
        self.new_notes_count = 0  # 已处理的新笔记数量
        self.excel_suffix = excel_suffix  # Excel 文件名后缀（如日期）

        self.browser = None
        self.context = None
        self.page = None
        self.playwright = None
        self.blogger_name = ""
        self.excel_file = os.path.join(self.output_dir, config.EXCEL_FILENAME)
        self.processed_titles: Set[str] = set()

    def _parse_manual_cookie(self) -> List[Dict]:
        """解析手动提供的Cookie字符串"""
        cookie_str = config.MANUAL_COOKIE.strip()
        if not cookie_str:
            return []

        cookies = []
        try:
            # 解析 "name=value; name2=value2" 格式
            for item in cookie_str.split(';'):
                item = item.strip()
                if not item:
                    continue
                if '=' in item:
                    name, value = item.split('=', 1)
                    cookies.append({
                        'name': name.strip(),
                        'value': value.strip(),
                        'domain': '.xiaohongshu.com',
                        'path': '/'
                    })
            return cookies
        except Exception as e:
            print(f"⚠️ 解析手动Cookie失败: {e}")
            return []

    def _load_manual_cookies(self) -> bool:
        """加载手动提供的Cookie"""
        cookies = self._parse_manual_cookie()
        if not cookies:
            return False

        try:
            if self.context:
                self.context.add_cookies(cookies)
                print(f"✓ 已加载手动提供的Cookie（{len(cookies)}个）")
                return True
        except Exception as e:
            print(f"⚠️ 加载手动Cookie失败: {e}")
        return False

    def _check_login_status(self, url: str = None) -> bool:
        """检查是否已登录"""
        try:
            # 访问主页并检查是否有登录相关的元素
            self._navigate_to_page(url or config.PROFILE_URL)
            time.sleep(2)

            # 检查是否有用户头像或用户名等登录标识
            # 小红书登录后会有特定的元素
            login_indicators = [
                '.user-info',
                '.avatar',
                '[class*="user"]',
                '.profile-info'
            ]

            for selector in login_indicators:
                try:
                    elem = self.page.query_selector(selector)
                    if elem:
                        return True
                except:
                    continue

            # 另一种检测方式：检查是否有登录按钮
            login_buttons = [
                '.login-btn',
                '.login-button',
                'button:has-text("登录")',
                'a:has-text("登录")'
            ]

            for selector in login_buttons:
                try:
                    elem = self.page.query_selector(selector)
                    if elem and elem.is_visible():
                        return False
                except:
                    continue

            # 如果找不到明确的登录或已登录标识，默认认为未登录
            return False
        except Exception as e:
            print(f"⚠️ 检查登录状态失败: {e}")
            return False
        
    def start(self) -> None:
        """启动浏览器"""
        self.playwright = sync_playwright().start()
        
        # 根据配置选择浏览器路径
        if config.BROWSER_TYPE.lower() == "edge":
            executable_path = config.EDGE_PATH
            print(f"使用 Edge 浏览器: {executable_path}")
        else:
            executable_path = config.CHROME_PATH
            print(f"使用 Chrome 浏览器: {executable_path}")
        if executable_path and not os.path.exists(executable_path):
            print(f"⚠️ 浏览器路径不存在，将使用 Playwright 默认浏览器: {executable_path}")
            executable_path = None
        
        self.browser = self.playwright.chromium.launch(
            headless=self.headless,
            executable_path=executable_path,
            args=config.BROWSER_ARGS
        )
        
        self.context = self.browser.new_context(
            viewport={'width': config.VIEWPORT_WIDTH, 'height': config.VIEWPORT_HEIGHT},
            user_agent=config.USER_AGENT,
            locale='zh-CN',
            timezone_id='Asia/Shanghai',
        )
        
        self._inject_anti_detection_script()
        
        self.page = self.context.new_page()
        os.makedirs(self.output_dir, exist_ok=True)
        
    def _inject_anti_detection_script(self) -> None:
        """注入反检测脚本"""
        self.context.add_init_script("""
            Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
            Object.defineProperty(navigator, 'plugins', { get: () => [1, 2, 3, 4, 5] });
            window.chrome = { runtime: {} };
        """)
        
    def close(self) -> None:
        """关闭浏览器"""
        if self.browser:
            self.browser.close()
        if self.playwright:
            self.playwright.stop()
            
    def get_all_notes(self, profile_url: str) -> Optional[List[NoteData]]:
        """
        获取用户所有博文

        流程：
        1. 启动浏览器
        2. 如果MANUAL_COOKIE不为空，尝试使用Cookie登录
        3. 如果Cookie登录失败或为空，进行手动登录
        4. 遍历所有帖子
        """
        try:
            self.start()
            print(f"正在访问: {profile_url}")

            # 检查是否有手动提供的Cookie
            if config.MANUAL_COOKIE.strip():
                print("\n检测到手动提供的Cookie，正在加载...")
                if self._load_manual_cookies():
                    print("正在验证Cookie登录状态...")
                    if self._check_login_status(profile_url):
                        print("✓ Cookie登录成功，无需手动登录")
                        return self._process_all_notes(profile_url)
                    else:
                        print("⚠️ Cookie登录失败，将使用手动登录")

            # 手动登录流程
            # 访问主页
            self._navigate_to_page(profile_url)
            print("✓ 已进入主页")

            # 获取第一篇笔记并点击（触发登录）
            first_note = self._get_first_note_element()
            if not first_note:
                print("⚠️ 未找到任何笔记")
                return None

            # 点击第一篇笔记触发登录弹窗
            print("\n正在点击第一篇笔记以触发登录...")
            self._click_element(first_note)
            print("✓ 已点击第一篇笔记，请完成登录")

            # 等待用户手动登录
            self._wait_for_login()

            # 登录完成后，重新获取所有笔记并遍历
            print("\n登录完成，开始遍历所有帖子...")
            return self._process_all_notes(profile_url)

        except Exception as e:
            print(f"获取失败: {e}")
            import traceback
            traceback.print_exc()
            return None
        finally:
            print("\n正在关闭浏览器...")
            self.close()
    
    def _navigate_to_page(self, url: str) -> None:
        """导航到指定页面"""
        self.page.goto(url, wait_until='networkidle', timeout=60000)
        time.sleep(config.PAGE_TRANSITION_DELAY)
        
    def _click_element(self, element: ElementHandle, force: bool = False) -> bool:
        """点击元素，支持多种点击方式"""
        try:
            element.scroll_into_view_if_needed()
            time.sleep(0.5)
            element.click(force=force)
            time.sleep(config.CLICK_DELAY)
            return True
        except Exception as e:
            print(f"普通点击失败: {e}")
            try:
                self.page.evaluate(
                    "(element) => { element.scrollIntoView({behavior: 'instant', block: 'center'}); element.click(); }",
                    element
                )
                time.sleep(config.CLICK_DELAY)
                return True
            except Exception as e2:
                print(f"JavaScript点击也失败: {e2}")
                return False
    
    def _wait_for_login(self) -> None:
        """等待用户完成登录"""
        print("\n" + "=" * 60)
        print("⚠️  请完成登录操作")
        print("=" * 60)
        print("如果看到登录弹窗/页面，请扫码或输入账号密码登录")
        print(f"等待{config.LOGIN_WAIT_TIME}秒...")
        time.sleep(config.LOGIN_WAIT_TIME)
        print("✓ 登录等待完成")
    
    def _get_first_note_element(self) -> Optional[ElementHandle]:
        """获取第一篇笔记的元素"""
        for selector in config.NOTE_SELECTORS:
            try:
                elements = self.page.query_selector_all(selector)
                if elements and len(elements) > 0:
                    print(f"✓ 使用选择器 '{selector}' 找到 {len(elements)} 个笔记")
                    return elements[0]
            except Exception:
                continue
        return None
    
    def _get_all_note_elements(self) -> List[ElementHandle]:
        """获取主页上所有笔记元素"""
        for selector in config.NOTE_SELECTORS:
            try:
                elements = self.page.query_selector_all(selector)
                if elements and len(elements) > 0:
                    return elements
            except Exception:
                continue
        return []
    
    def _get_note_links(self, elements: List[ElementHandle]) -> List[str]:
        """从笔记元素中提取链接"""
        links = []
        for elem in elements:
            try:
                href = elem.get_attribute('href')
                if href and ('/explore/' in href or '/user/profile/' in href):
                    full_url = href if href.startswith('http') else f"https://www.xiaohongshu.com{href}"
                    if full_url not in links:
                        links.append(full_url)
            except Exception:
                continue
        return links
    
    def _scroll_to_load_all_notes(self) -> None:
        """滚动页面加载所有帖子"""
        max_scroll_attempts = config.SCROLL_MAX_ATTEMPTS
        scroll_pause_time = config.SCROLL_PAUSE_TIME
        no_change_count = 0
        last_note_count = 0
        
        print(f"  开始滚动加载...")
        
        for attempt in range(max_scroll_attempts):
            # 获取当前帖子数量
            current_elements = self._get_all_note_elements()
            current_count = len(current_elements)
            
            # 每5次滚动输出一次进度
            if (attempt + 1) % 5 == 0 or current_count != last_note_count:
                print(f"  第 {attempt + 1} 次滚动，当前帖子数: {current_count}")
            
            # 检查是否有新帖子加载
            if current_count == last_note_count:
                no_change_count += 1
                # 连续无变化次数达到阈值，认为已加载完毕
                if no_change_count >= config.SCROLL_NO_CHANGE_THRESHOLD:
                    print(f"✓ 滚动完成，共加载 {current_count} 篇帖子")
                    break
            else:
                no_change_count = 0
                last_note_count = current_count
            
            # 模拟更真实的滚动：每次滚动一屏的高度
            scroll_height = self.page.evaluate("window.innerHeight")
            current_scroll = self.page.evaluate("window.scrollY")
            new_scroll = current_scroll + scroll_height * 0.8
            
            self.page.evaluate(f"window.scrollTo(0, {new_scroll})")
            time.sleep(scroll_pause_time)
        
        if no_change_count < config.SCROLL_NO_CHANGE_THRESHOLD:
            print(f"✓ 达到最大滚动次数，共加载 {last_note_count} 篇帖子")
    
    def _init_excel(self) -> None:
        """初始化Excel文件"""
        if not os.path.exists(self.excel_file):
            wb = Workbook()
            ws = wb.active
            ws.title = "笔记索引"
            ws.append(config.EXCEL_HEADERS)
            # 设置表头样式
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            wb.save(self.excel_file)
            print(f"✓ 创建Excel索引文件: {self.excel_file}")
        else:
            # 加载已处理的标题
            self._load_processed_titles()
    
    def _load_processed_titles(self) -> None:
        """从Excel加载已处理的标题"""
        try:
            wb = load_workbook(self.excel_file)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[1]:  # 标题列
                    self.processed_titles.add(row[1])
            print(f"✓ 从Excel加载了 {len(self.processed_titles)} 个已处理的笔记")
        except Exception as e:
            print(f"⚠️ 加载Excel失败: {e}")
    
    def _add_note_to_excel(self, index: int, title: str, content: str, likes: str, 
                           collects: str, comments: str, note_id: str, 
                           url: str, status: str = "待处理") -> bool:
        """添加笔记到Excel，如果标题已存在则返回False"""
        if title in self.processed_titles:
            return False
        
        try:
            wb = load_workbook(self.excel_file)
            ws = wb.active
            ws.append([index, title, content, likes, collects, comments, note_id, url, status])
            wb.save(self.excel_file)
            self.processed_titles.add(title)
            return True
        except Exception as e:
            print(f"⚠️ 添加笔记到Excel失败: {e}")
            return False
    
    def _update_note_status(self, title: str, status: str, content: str = "",
                           likes: str = "", collects: str = "", comments: str = "",
                           new_title: str = "") -> None:
        """更新笔记状态和详细信息"""
        try:
            wb = load_workbook(self.excel_file)
            ws = wb.active
            for row in ws.iter_rows(min_row=2):
                if row[1].value == title:
                    # 如果提供了新标题，更新标题
                    if new_title and new_title != title:
                        row[1].value = new_title
                        # 更新已处理标题集合
                        self.processed_titles.discard(title)
                        self.processed_titles.add(new_title)
                        print(f"  ✓ 更新标题: {title[:30]}... -> {new_title[:30]}...")
                    row[8].value = status  # 状态列
                    # 如果提供了其他信息，也更新
                    if content:
                        row[2].value = content
                    if likes:
                        row[3].value = likes
                    if collects:
                        row[4].value = collects
                    if comments:
                        row[5].value = comments
                    wb.save(self.excel_file)
                    break
        except Exception as e:
            print(f"⚠️ 更新笔记状态失败: {e}")

    def _sanitize_filename(self, filename: str) -> str:
        """清理Windows文件名中的非法字符"""
        filename = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", filename).strip()
        filename = re.sub(r"\s+", " ", filename)
        return filename[:80].strip(" .") or "notes_index"

    def _extract_blogger_name(self) -> str:
        """从博主主页提取博主昵称，用于命名Excel文件"""
        selectors = [
            ".user-nickname .user-name",
            ".user-name",
            "[class*='user-nickname'] [class*='user-name']",
            "[class*='user-name']",
            ".nickname",
            "[class*='nickname']",
        ]

        for selector in selectors:
            try:
                elem = self.page.query_selector(selector)
                if elem:
                    name = elem.inner_text().strip()
                    if name:
                        return " ".join(name.split())
            except Exception:
                continue

        return ""

    def _configure_excel_for_blogger(self) -> None:
        """进入主页后，先提取博主名，并用博主名命名Excel"""
        blogger_name = self._extract_blogger_name()
        if not blogger_name:
            print("⚠️ 未提取到博主名，继续使用默认Excel文件名")
            return

        self.blogger_name = blogger_name
        safe_name = self._sanitize_filename(blogger_name)
        
        # 如果有后缀（如日期），添加到文件名
        if self.excel_suffix:
            self.excel_file = os.path.join(self.output_dir, f"{safe_name}_{self.excel_suffix}.xlsx")
        else:
            self.excel_file = os.path.join(self.output_dir, f"{safe_name}.xlsx")
        
        print(f"✓ 提取到博主名: {blogger_name}")
        print(f"✓ Excel文件将保存为: {self.excel_file}")
    
    def _get_note_info_from_element(self, element) -> Optional[Dict]:
        """从元素中提取笔记信息（标题和链接）"""
        try:
            # 获取链接
            href = element.get_attribute('href')
            if not href:
                return None
            
            full_url = href if href.startswith('http') else f"https://www.xiaohongshu.com{href}"
            note_id = self._extract_note_id(full_url)
            
            # 尝试多种方式获取标题
            title = ""
            
            # 方法1: 从img的alt属性获取
            if not title:
                try:
                    img = element.query_selector('img')
                    if img:
                        title = img.get_attribute('alt') or ""
                except:
                    pass
            
            # 方法2: 从title属性获取
            if not title:
                try:
                    title = element.get_attribute('title') or ""
                except:
                    pass
            
            # 方法3: 从class包含title的元素获取
            if not title:
                try:
                    title_elem = element.query_selector('.title, [class*="title"], .desc, [class*="desc"]')
                    if title_elem:
                        title = title_elem.inner_text().strip()
                except:
                    pass
            
            # 方法4: 从所有子元素的文本内容获取（排除特定元素）
            if not title:
                try:
                    # 获取元素内的文本，但排除一些非标题元素
                    all_text = element.inner_text().strip()
                    # 分割成行，取第一行非空内容
                    lines = [line.strip() for line in all_text.split('\n') if line.strip()]
                    if lines:
                        # 过滤掉纯数字（可能是点赞数等）
                        for line in lines[:3]:  # 只看前3行
                            if not line.isdigit() and len(line) > 1:
                                title = line
                                break
                except:
                    pass
            
            # 方法5: 从data属性获取
            if not title:
                try:
                    title = element.get_attribute('data-title') or ""
                except:
                    pass
            
            # 清理标题
            if title:
                # 移除多余空白
                title = ' '.join(title.split())
                # 限制长度
                if len(title) > 100:
                    title = title[:100] + "..."
            
            if not title:
                title = f"笔记_{note_id[:8]}"
            
            return {
                'note_id': note_id,
                'title': title,
                'url': full_url
            }
        except Exception as e:
            return None
    
    def _scroll_and_collect_notes(self, profile_url: str) -> List[Dict]:
        """滚动页面并收集所有笔记信息（旧方法，保留兼容）"""
        return self._scroll_and_collect_notes_incremental(profile_url)

    def _scroll_and_collect_notes_incremental(self, profile_url: str) -> List[Dict]:
        """滚动页面并收集笔记信息（支持增量爬取）
        
        如果设置了 existing_note_ids，则会：
        1. 跳过已存在的 note_id
        2. 只收集新的笔记，直到达到 max_notes 数量
        """
        is_incremental = len(self.existing_note_ids) > 0
        if is_incremental:
            print(f"\n[增量爬取] 开始滚动收集新笔记（目标：{self.max_notes} 条新笔记）...")
            print(f"[增量爬取] 已存在 {len(self.existing_note_ids)} 条笔记，将自动跳过")
        else:
            print("\n开始滚动收集所有笔记信息...")
        
        all_notes_info = []
        seen_titles = set()
        seen_note_ids = set()
        max_scroll_attempts = config.SCROLL_MAX_ATTEMPTS
        scroll_pause_time = config.SCROLL_PAUSE_TIME
        no_new_count = 0
        skipped_count = 0
        
        for attempt in range(max_scroll_attempts):
            # 检查是否已收集足够的新笔记
            if self.max_notes and len(all_notes_info) >= self.max_notes:
                print(f"✓ 已收集足够的新笔记（{len(all_notes_info)} 条），停止滚动")
                break
            
            # 获取当前页面所有笔记元素
            note_elements = self._get_all_note_elements()
            current_new_count = 0
            current_skipped = 0
            
            for elem in note_elements:
                info = self._get_note_info_from_element(elem)
                if not info:
                    continue
                    
                # 检查标题是否已见过
                if info['title'] in seen_titles:
                    continue
                seen_titles.add(info['title'])
                
                # 检查 note_id 是否已存在（增量爬取逻辑）
                note_id = info.get('note_id', '')
                if note_id:
                    # 如果已经在这个批次中见过，跳过
                    if note_id in seen_note_ids:
                        continue
                    seen_note_ids.add(note_id)
                    
                    # 如果已在数据库中存在，跳过
                    if note_id in self.existing_note_ids:
                        skipped_count += 1
                        current_skipped += 1
                        continue
                
                # 是新笔记，添加到列表
                all_notes_info.append(info)
                current_new_count += 1
                
                # 添加到Excel（检查重复）
                self._add_note_to_excel(
                    len(all_notes_info), 
                    info['title'], 
                    "",  # 内容稍后获取
                    "",  # 点赞数稍后获取
                    "",  # 收藏数稍后获取
                    "",  # 评论数稍后获取
                    info['note_id'], 
                    info['url']
                )
                
                # 检查是否已收集足够
                if self.max_notes and len(all_notes_info) >= self.max_notes:
                    break
            
            # 每10次滚动输出进度
            if (attempt + 1) % 10 == 0 or current_new_count > 0 or current_skipped > 0:
                progress_msg = f"  第 {attempt + 1} 次滚动，新笔记：{len(all_notes_info)} 条"
                if is_incremental:
                    progress_msg += f"（本次新增 {current_new_count} 条，跳过 {current_skipped} 条已存在）"
                    progress_msg += f"，累计跳过：{skipped_count} 条"
                else:
                    progress_msg += f"（本次新增 {current_new_count} 条）"
                print(progress_msg)
            
            # 检查是否还有新内容
            if current_new_count == 0 and current_skipped == 0:
                no_new_count += 1
                if no_new_count >= config.SCROLL_NO_CHANGE_THRESHOLD:
                    if is_incremental:
                        print(f"✓ 滚动收集完成，共收集 {len(all_notes_info)} 条新笔记，跳过 {skipped_count} 条已存在")
                    else:
                        print(f"✓ 滚动收集完成，共 {len(all_notes_info)} 篇笔记")
                    break
            else:
                no_new_count = 0
            
            # 滚动页面
            scroll_height = self.page.evaluate("window.innerHeight")
            current_scroll = self.page.evaluate("window.scrollY")
            new_scroll = current_scroll + scroll_height * 0.8
            
            self.page.evaluate(f"window.scrollTo(0, {new_scroll})")
            time.sleep(scroll_pause_time)
        
        return all_notes_info
    
    def _process_all_notes(self, profile_url: str) -> List[NoteData]:
        """处理所有笔记"""
        all_notes = []
        
        # 返回主页
        self._navigate_to_page(profile_url)

        # 先提取博主昵称，用博主名命名本次Excel文件
        self._configure_excel_for_blogger()
        
        # 初始化Excel
        self._init_excel()
        
        # 滚动收集所有笔记信息到Excel（带增量爬取逻辑）
        notes_info = self._scroll_and_collect_notes_incremental(profile_url)
        
        if not notes_info:
            print("⚠️ 未收集到任何新笔记信息（可能都已存在）")
            return all_notes

        print(f"\n✓ 共收集到 {len(notes_info)} 篇新笔记，开始处理...")
        
        # 重新滚动到顶部
        print("滚动回顶部，开始处理...")
        self.page.evaluate("window.scrollTo(0, 0)")
        time.sleep(1)
        
        # 根据Excel记录遍历处理每篇笔记
        for i, note_info in enumerate(notes_info):
            title = note_info['title']
            
            # 检查是否已处理过
            if title in self.processed_titles and i < len(self.processed_titles):
                # 检查Excel中的状态
                try:
                    wb = load_workbook(self.excel_file)
                    ws = wb.active
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if row[1] == title and row[8] == "已完成":
                            print(f"  跳过已处理: {title[:30]}...")
                            continue
                except:
                    pass
            
            print(f"\n{'=' * 60}")
            print(f"正在处理第 {i+1}/{len(notes_info)} 篇笔记")
            print(f"标题: {title[:50]}...")
            print(f"{'=' * 60}")
            
            # 在主页找到并点击该帖子
            note_elements = self._get_all_note_elements()
            target_element = None
            
            for elem in note_elements:
                info = self._get_note_info_from_element(elem)
                if info and info['title'] == title:
                    target_element = elem
                    break
            
            if not target_element:
                print(f"⚠️ 未找到帖子: {title[:30]}...，跳过")
                continue
            
            # 滚动到元素可见
            target_element.scroll_into_view_if_needed()
            time.sleep(0.5)
            
            # 点击帖子
            print(f"正在点击帖子...")
            if not self._click_element(target_element):
                print(f"⚠️ 点击帖子失败，跳过")
                continue
            
            # 获取帖子详情
            note_data = self._extract_note_detail(note_info['url'], i + 1)
            if note_data:
                all_notes.append(note_data)
                self._save_note(note_data, i + 1)
                # 更新Excel中的完整信息（包括准确的标题）
                self._update_note_status(
                    title,
                    "已完成",
                    content=note_data.content,
                    likes=note_data.likes,
                    collects=note_data.collects,
                    comments=note_data.comments,
                    new_title=note_data.title  # 传入准确的标题
                )
            else:
                self._update_note_status(title, "处理失败")
            
            # 关闭详情页（如果不是最后一个帖子）
            if i < len(notes_info) - 1:
                if not self._close_note_detail():
                    print("⚠️ 关闭详情页失败，返回主页重新获取...")
                    self._navigate_to_page(profile_url)
        
        return all_notes
    
    def _close_note_detail(self) -> bool:
        """关闭笔记详情页"""
        for selector in config.CLOSE_SELECTORS:
            try:
                close_btn = self.page.query_selector(selector)
                if close_btn:
                    print(f"  找到关闭按钮: {selector}")
                    if self._click_element(close_btn):
                        print("  ✓ 已关闭详情页")
                        time.sleep(1)
                        return True
            except Exception:
                continue
        
        # 尝试按 ESC 键关闭
        try:
            self.page.keyboard.press('Escape')
            print("  尝试按 ESC 键关闭")
            time.sleep(1)
            return True
        except Exception:
            pass
        
        return False
    
    def _extract_note_detail(self, note_url: str, index: int) -> Optional[NoteData]:
        """从当前页面提取笔记详情"""
        try:
            # 使用当前页面的实际URL（更准确）
            current_url = self.page.url
            print(f"当前URL: {current_url}")
            
            # 等待页面加载
            time.sleep(config.DETAIL_PAGE_WAIT)
            
            # 从当前URL提取笔记ID
            note_id = self._extract_note_id(current_url)
            title = self._extract_text(config.TITLE_SELECTORS)
            content = self._extract_content()
            
            # 检查是否为视频帖子
            is_video = self._is_video_post()
            videos = []
            images = []
            
            if is_video:
                # 提取视频地址
                videos = self._extract_videos()
                print(f"  检测到视频帖子，视频数量: {len(videos)}")
            else:
                # 提取图片
                images = self._extract_article_images()
            
            stats = self._extract_stats()
            
            note_data = NoteData(
                index=index,
                note_id=note_id,
                title=title,
                content=content,
                images=images,
                videos=videos,
                url=current_url,
                likes=stats.get('likes', ''),
                collects=stats.get('collects', ''),
                comments=stats.get('comments', ''),
                is_video=is_video
            )
            
            print(f"✓ 成功获取笔记: {title[:50] if title else '无标题'}...")
            if is_video:
                print(f"  - 视频数量: {len(videos)}")
            else:
                print(f"  - 图片数量: {len(images)}")
            print(f"  - 内容长度: {len(content)} 字符")
            print(f"  - 点赞: {stats.get('likes', '0')} | 收藏: {stats.get('collects', '0')} | 评论: {stats.get('comments', '0')}")
            
            return note_data
            
        except Exception as e:
            print(f"✗ 获取笔记详情失败: {e}")
            return None
    
    def _extract_note_id(self, url: str) -> str:
        """从URL中提取笔记ID"""
        patterns = [
            r'/explore/([a-zA-Z0-9]+)',
            r'/user/profile/[^/]+/([a-zA-Z0-9]+)',
        ]
        for pattern in patterns:
            match = re.search(pattern, url)
            if match:
                return match.group(1)
        return ""
    
    def _extract_text(self, selectors: List[str]) -> str:
        """使用选择器列表提取文本"""
        for selector in selectors:
            try:
                elem = self.page.query_selector(selector)
                if elem:
                    text = elem.inner_text().strip()
                    if text:
                        # 如果是title标签，清理后缀
                        if selector == 'title':
                            text = self._clean_title(text)
                        return text
            except Exception:
                continue
        return ""
    
    def _clean_title(self, title: str) -> str:
        """清理标题，移除小红书后缀"""
        # 移除常见的后缀
        suffixes = [' - 小红书', ' | 小红书', ' — 小红书']
        for suffix in suffixes:
            if title.endswith(suffix):
                return title[:-len(suffix)].strip()
        return title
    
    def _extract_content(self) -> str:
        """提取笔记正文内容"""
        # 首先尝试常规选择器
        content = self._extract_text(config.CONTENT_SELECTORS)
        if content:
            return content
        
        # 尝试从script标签中获取
        try:
            scripts = self.page.query_selector_all('script')
            for script in scripts:
                text = script.inner_text()
                if 'desc' in text or 'content' in text:
                    matches = re.findall(r'"desc":"([^"]+)"', text)
                    if matches:
                        return matches[0].replace('\\n', '\n')
        except Exception:
            pass
        
        return ""
    
    def _is_video_post(self) -> bool:
        """检查当前帖子是否为视频"""
        try:
            # 检查是否有video元素
            video_elem = self.page.query_selector('video')
            if video_elem:
                return True

            # 检查是否有视频相关的标识
            video_indicators = [
                '.video-container',
                '.player-container',
                '[class*="video"]',
                'video[src^="blob:"]'
            ]
            for selector in video_indicators:
                try:
                    elem = self.page.query_selector(selector)
                    if elem:
                        return True
                except:
                    continue
        except:
            pass
        return False

    def _extract_videos(self) -> List[str]:
        """
        从小红书作品页面提取无水印视频直链
        核心逻辑：通过 JavaScript 在浏览器内提取视频URL，避免序列化问题
        """
        videos = []

        try:
            # 方法1: 在浏览器内提取视频URL（避免大数据序列化问题）
            try:
                video_url = self.page.evaluate("""() => {
                    const state = window.__INITIAL_STATE__;
                    if (!state) return null;
                    
                    let url = null;
                    
                    // 路径1：noteDetail -> note -> video -> playAddr
                    if (state.noteDetail && state.noteDetail.note && state.noteDetail.note.video) {
                        const video = state.noteDetail.note.video;
                        if (video.playAddr) url = video.playAddr;
                        else if (video.playUrl) url = video.playUrl;
                    }
                    
                    // 路径2：feed -> notes 结构
                    if (!url && state.feed && state.feed.notes) {
                        const notes = state.feed.notes;
                        for (let key in notes) {
                            if (notes[key].video) {
                                const video = notes[key].video;
                                if (video.playAddr) {
                                    url = video.playAddr;
                                    break;
                                } else if (video.playUrl) {
                                    url = video.playUrl;
                                    break;
                                }
                            }
                        }
                    }
                    
                    // 处理URL转义
                    if (url) {
                        url = url.replace(/\\\\/g, '').replace(/u0026/g, '&');
                    }
                    
                    return url;
                }""")
                
                if video_url:
                    videos.append(video_url)
                    print(f"  从 JavaScript 提取到视频地址")
                    print(f"    视频: {video_url[:80]}...")
                    return videos
                else:
                    print("  JavaScript 未找到视频地址")
                        
            except Exception as e:
                print(f"  JavaScript 获取失败: {e}")

            # 方法2: 从页面源码正则提取（备用）
            print("  尝试从页面源码提取...")
            page_content = self.page.content()
            
            # 查找 window.__INITIAL_STATE__ - 使用更精确的正则
            state_pattern = re.compile(r'window\.__INITIAL_STATE__\s*=\s*({.+?});\s*</script>', re.DOTALL)
            state_match = state_pattern.search(page_content)
            
            if state_match:
                state_json_str = state_match.group(1)
                # 处理转义字符
                state_json_str = state_json_str.replace('\\u0026', '&').replace('\\n', '').replace('\\t', '')
                
                try:
                    state_data = json.loads(state_json_str)
                    
                    video_url = None
                    if "noteDetail" in state_data and "note" in state_data["noteDetail"]:
                        note_data = state_data["noteDetail"]["note"]
                        if "video" in note_data:
                            video_info = note_data["video"]
                            if "playAddr" in video_info:
                                video_url = video_info["playAddr"]
                            elif "playUrl" in video_info:
                                video_url = video_info["playUrl"]
                    
                    if video_url:
                        video_url = video_url.replace("\\", "").replace("u0026", "&")
                        videos.append(video_url)
                        print(f"  从页面源码提取到视频地址")
                        print(f"    视频: {video_url[:80]}...")
                        return videos
                        
                except json.JSONDecodeError as e:
                    print(f"  JSON解析失败: {e}")
            else:
                print("  未找到 window.__INITIAL_STATE__")
            
            # 方法3: 兜底正则匹配 - 直接从页面中查找
            print("  尝试兜底正则匹配...")
            backup_patterns = [
                r'"playAddr":"(https?://[^"]+)"',
                r'"playUrl":"(https?://[^"]+)"',
                r'(https?://sns-video[^\s"\'<>]+\.mp4[^\s"\'<>]*)',
            ]
            
            for pattern in backup_patterns:
                backup_match = re.search(pattern, page_content)
                if backup_match:
                    video_url = backup_match.group(1)
                    if video_url:
                        video_url = video_url.replace("\\", "").replace("u0026", "&")
                        videos.append(video_url)
                        print(f"  通过兜底正则提取到视频地址")
                        print(f"    视频: {video_url[:80]}...")
                        return videos
            
            print("  未提取到视频下载地址")
            
        except Exception as e:
            print(f"  提取视频地址失败: {e}")

        return videos
    
    def _extract_article_images(self) -> List[str]:
        """提取文章图片（排除评论图片、头像和视频）"""
        # 首先检查是否是视频帖子
        if self._is_video_post():
            print("  检测到视频帖子，跳过图片提取")
            return []
        
        images = []
        
        # 更精确的图片选择器，限制在当前笔记详情页的主内容区域
        article_image_selectors = [
            # 主内容区域的图片
            '.note-content .img-container img',
            '.media-container .img-container img',
            '.main-content img',
            '.note-detail .img-container img',
            # 轮播图
            '.swiper-wrapper .swiper-slide img',
            '.note-slider-img img',
            # 详情页主图
            'div[class*="note"] img[src*="xhscdn.com"]',
        ]
        
        for selector in article_image_selectors:
            try:
                imgs = self.page.query_selector_all(selector)
                print(f"  选择器 '{selector}' 找到 {len(imgs)} 个图片")
                
                for img in imgs:
                    src = img.get_attribute('src') or img.get_attribute('data-src')
                    if src and 'xhscdn.com' in src:
                        # 过滤掉头像和小图标
                        if self._is_article_image(src):
                            src = src.strip().strip('`').strip()
                            # 确保不是blob URL（视频）
                            if not src.startswith('blob:') and src not in images:
                                images.append(src)
                
                # 如果找到足够图片，不再继续
                if len(images) >= 1:
                    break
                    
            except Exception as e:
                print(f"  选择器 '{selector}' 出错: {e}")
                continue
        
        # 去重并限制数量
        unique_images = list(dict.fromkeys(images))  # 保持顺序去重
        
        print(f"  共提取到 {len(unique_images)} 张文章图片")
        return unique_images
    
    def _is_article_image(self, img_url: str) -> bool:
        """判断是否为文章图片（排除头像、评论图等）"""
        # 排除常见的头像/评论图片特征
        excluded_patterns = [
            r'avatar',
            r'comment',
            r'user',
            r'icon',
            r'emoji',
            r'webp/sd',
        ]
        
        img_url_lower = img_url.lower()
        for pattern in excluded_patterns:
            if re.search(pattern, img_url_lower):
                return False
        
        return True
    
    def _extract_stats(self) -> Dict[str, str]:
        """提取笔记统计数据（点赞、收藏、评论）"""
        stats = {'likes': '', 'collects': '', 'comments': ''}
        
        # 方法1: 根据HTML结构，从 .left 容器中提取
        try:
            left_container = self.page.query_selector('.left')
            if left_container:
                # 点赞数
                like_elem = left_container.query_selector('.like-wrapper .count')
                if like_elem:
                    stats['likes'] = like_elem.inner_text().strip()
                
                # 收藏数
                collect_elem = left_container.query_selector('.collect-wrapper .count')
                if collect_elem:
                    stats['collects'] = collect_elem.inner_text().strip()
                
                # 评论数
                comment_elem = left_container.query_selector('.chat-wrapper .count')
                if comment_elem:
                    stats['comments'] = comment_elem.inner_text().strip()
                
                if any(stats.values()):
                    return stats
        except Exception:
            pass
        
        # 方法2: 使用配置的选择器
        for stat_type, selectors in config.STATS_SELECTORS.items():
            for selector in selectors:
                try:
                    elem = self.page.query_selector(selector)
                    if elem:
                        text = elem.inner_text().strip()
                        if text:
                            stats[stat_type] = text
                            break
                except Exception:
                    continue
        
        return stats
    
    def _save_note(self, note: NoteData, index: int) -> None:
        """保存笔记信息（仅更新Excel，不创建文件夹）"""
        # 只保存到Excel，不创建文件夹和下载图片
        print(f"\n笔记信息已保存到Excel: {note.title[:50] if note.title else '无标题'}...")
        print(f"  - 图片数量: {len(note.images)}")
        print(f"  - 内容长度: {len(note.content)} 字符")
        print(f"  - 点赞: {note.likes} | 收藏: {note.collects} | 评论: {note.comments}")
        print(f"✓ 笔记处理完成（仅Excel）")
    
    def _download_images(self, image_urls: List[str], folder_path: str) -> List[str]:
        """下载图片"""
        downloaded = []

        if not image_urls:
            return downloaded

        print(f"正在下载 {len(image_urls)} 张图片...")

        for i, img_url in enumerate(image_urls):
            try:
                img_url = img_url.strip().strip('`').strip()

                filename = self._download_image_with_browser(img_url, folder_path, i + 1)
                if filename:
                    downloaded.append(filename)

            except Exception as e:
                print(f"  ✗ 下载失败: {e}")

        return downloaded

    def _download_videos(self, video_urls: List[str], folder_path: str) -> List[str]:
        """下载视频"""
        downloaded = []

        if not video_urls:
            return downloaded

        print(f"正在下载 {len(video_urls)} 个视频...")

        for i, video_url in enumerate(video_urls):
            try:
                video_url = video_url.strip().strip('`').strip()

                filename = self._download_video_with_browser(video_url, folder_path, i + 1)
                if filename:
                    downloaded.append(filename)

            except Exception as e:
                print(f"  ✗ 视频下载失败: {e}")

        return downloaded
    
    def _download_image_with_browser(self, img_url: str, folder_path: str, index: int) -> Optional[str]:
        """使用浏览器下载图片（携带登录态）"""
        try:
            result = self.page.evaluate("""async (url) => {
                try {
                    const response = await fetch(url, {
                        method: 'GET',
                        credentials: 'include',
                        headers: {
                            'Accept': 'image/webp,image/apng,image/*,*/*;q=0.8'
                        }
                    });

                    if (response.ok) {
                        const blob = await response.blob();
                        const reader = new FileReader();

                        return new Promise((resolve) => {
                            reader.onloadend = () => {
                                const base64data = reader.result.split(',')[1];
                                resolve({
                                    success: true,
                                    data: base64data,
                                    type: blob.type
                                });
                            };
                            reader.readAsDataURL(blob);
                        });
                    }
                    return { success: false, status: response.status };
                } catch (e) {
                    return { success: false, error: e.message };
                }
            }""", img_url)

            if result and result.get('success'):
                image_data = base64.b64decode(result['data'])
                ext = self._get_image_extension(result.get('type', ''))
                filename = f"image_{index:02d}.{ext}"
                filepath = os.path.join(folder_path, filename)

                with open(filepath, 'wb') as f:
                    f.write(image_data)

                if len(image_data) > 100:
                    print(f"  ✓ 下载: {filename} ({len(image_data)} bytes)")
                    return filename

        except Exception as e:
            print(f"  ✗ 下载失败: {e}")

        return None

    def _download_video_with_browser(self, video_url: str, folder_path: str, index: int) -> Optional[str]:
        """使用浏览器下载视频（携带登录态和完整请求头）"""
        try:
            print(f"  开始下载视频 {index}: {video_url[:60]}...")

            result = self.page.evaluate("""async (url) => {
                try {
                    const response = await fetch(url, {
                        method: 'GET',
                        credentials: 'include',
                        headers: {
                            'Accept': 'video/mp4,video/*,*/*;q=0.9',
                            'Accept-Language': 'zh-CN,zh;q=0.9',
                            'Referer': 'https://www.xiaohongshu.com/',
                            'Origin': 'https://www.xiaohongshu.com'
                        }
                    });

                    if (!response.ok) {
                        return { success: false, status: response.status, error: `HTTP ${response.status}` };
                    }

                    const blob = await response.blob();
                    const arrayBuffer = await blob.arrayBuffer();
                    const uint8Array = new Uint8Array(arrayBuffer);

                    return {
                        success: true,
                        data: Array.from(uint8Array),
                        type: blob.type,
                        size: blob.size
                    };
                } catch (e) {
                    return { success: false, error: e.message };
                }
            }""", video_url)

            if result and result.get('success'):
                video_data = bytes(result['data'])
                size_mb = len(video_data) / (1024 * 1024)

                if size_mb < 0.001:
                    print(f"  ✗ 视频文件太小，可能下载失败: {len(video_data)} bytes")
                    return None

                ext = 'mp4'
                filename = f"video_{index:02d}.{ext}"
                filepath = os.path.join(folder_path, filename)

                with open(filepath, 'wb') as f:
                    f.write(video_data)

                print(f"  ✓ 下载视频: {filename} ({size_mb:.2f} MB)")
                return filename
            else:
                error_msg = result.get('error', f"HTTP {result.get('status', 'unknown')}")
                print(f"  ✗ 视频下载失败: {error_msg}")

        except Exception as e:
            print(f"  ✗ 视频下载异常: {e}")

        return None
    
    def _get_image_extension(self, content_type: str) -> str:
        """根据Content-Type获取图片扩展名"""
        if 'webp' in content_type:
            return 'webp'
        elif 'png' in content_type:
            return 'png'
        elif 'jpeg' in content_type or 'jpg' in content_type:
            return 'jpg'
        return 'jpg'
    
    def _generate_markdown(self, note: NoteData, image_files: List[str]) -> str:
        """生成Markdown内容（标题、正文、统计数据、视频/图片）"""
        lines = []

        # 标题
        if note.title:
            lines.append(f"# {note.title}")
            lines.append("")

        # 正文内容
        if note.content:
            lines.append(note.content)
            lines.append("")

        # 视频或图片信息
        if note.is_video:
            lines.append("---")
            lines.append("📹 **视频帖子**")
            if note.videos:
                lines.append(f"视频数量: {len(note.videos)}")
                lines.append("")
                lines.append("视频链接:")
                for i, video_url in enumerate(note.videos, 1):
                    lines.append(f"{i}. [{video_url[:60]}...]({video_url})")
            else:
                lines.append("*未提取到视频地址*")
            lines.append("---")
            lines.append("")

        # 统计数据（点赞、收藏、评论）
        stats_parts = []
        if note.likes:
            stats_parts.append(f"点赞: {note.likes}")
        if note.collects:
            stats_parts.append(f"收藏: {note.collects}")
        if note.comments:
            stats_parts.append(f"评论: {note.comments}")

        if stats_parts:
            lines.append("---")
            lines.append(" | ".join(stats_parts))
            lines.append("---")
            lines.append("")

        # 原文链接
        lines.append(f"[原文链接]({note.url})")
        lines.append("")

        return "\n".join(lines)
    
    def _write_file(self, filepath: str, content: str) -> None:
        """写入文本文件"""
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(content)
    
    def _write_json(self, filepath: str, data: Dict) -> None:
        """写入JSON文件"""
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)


def main():
    """
    主函数
    所有配置请在 config.py 中修改
    """
    print("=" * 60)
    print("小红书博主博文获取工具")
    print("=" * 60)
    print(f"\n配置:")
    print(f"  - 下载图片: {'是' if config.DOWNLOAD_IMAGES else '否'}")
    print(f"  - 目标博主: {config.PROFILE_URL}")
    print(f"  - 输出目录: {config.OUTPUT_DIR}")
    print(f"  - 无头模式: {'是' if config.HEADLESS else '否'}")
    print("\n执行流程:")
    print("1. 进入主页")
    print("2. 点击第一篇博文触发登录")
    print("3. 等待用户手动登录")
    print("4. 登录后滚动收集所有帖子信息到Excel")
    print("5. 根据Excel记录遍历处理每篇帖子")
    print("6. 每篇博文保存到独立文件夹（markdown + 可选图片）")
    print("=" * 60)
    
    scraper = XiaoHongShuScraper()
    notes = scraper.get_all_notes(config.PROFILE_URL)
    
    if not notes:
        print("\n✗ 未能获取到博文数据")
        return
    
    print("\n" + "=" * 60)
    print(f"📄 共成功获取 {len(notes)} 篇博文")
    print(f"📁 所有文件已保存到: {scraper.output_dir}")
    print("=" * 60)
    
    for i, note in enumerate(notes):
        print(f"\n第 {i+1} 篇博文")
        print(f"  🆔 笔记ID: {note.note_id or 'N/A'}")
        print(f"  📝 标题: {note.title[:50] if note.title else '无标题'}...")
        print(f"  🖼️ 图片数量: {len(note.images)}")
        print(f"  👍 点赞: {note.likes or '0'} | ⭐ 收藏: {note.collects or '0'} | 💬 评论: {note.comments or '0'}")
    
    print("\n" + "=" * 60)
    print("✓ 所有任务完成！")
    print("=" * 60)
    
    return notes


if __name__ == "__main__":
    main()
