"""Read Xiaohongshu crawler Excel files for the web app."""

import argparse
import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook


TITLE_HEADERS = {"标题", "title", "Title"}
CONTENT_HEADERS = {"正文", "内容", "文案", "content", "Content"}
NOTE_ID_HEADERS = {"笔记ID", "note_id", "noteId", "Note ID"}
URL_HEADERS = {"链接", "URL", "url", "Url"}


def _safe_text(value):
    if value is None:
        return ""
    return str(value).strip()


def _iter_excel_files(root):
    root = Path(root).resolve()
    if not root.exists():
        return []

    return [
        path
        for path in root.rglob("*.xlsx")
        if path.is_file() and not path.name.startswith("~$")
    ]


def _find_columns(headers):
    title_col = None
    content_col = None
    note_id_col = None
    url_col = None

    for index, header in enumerate(headers):
        text = _safe_text(header)
        if text in TITLE_HEADERS:
            title_col = index
        elif text in CONTENT_HEADERS:
            content_col = index
        elif text in NOTE_ID_HEADERS:
            note_id_col = index
        elif text in URL_HEADERS:
            url_col = index

    return title_col, content_col, note_id_col, url_col


def _read_posts(path, limit=None):
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = next(rows, [])
        title_col, content_col, note_id_col, url_col = _find_columns(headers)

        if title_col is None or content_col is None:
            return []

        posts = []
        for row in rows:
            title = _safe_text(row[title_col] if title_col < len(row) else "")
            content = _safe_text(row[content_col] if content_col < len(row) else "")
            if not title or not content:
                continue

            posts.append({
                "title": title,
                "content": content,
                "noteId": _safe_text(row[note_id_col] if note_id_col is not None and note_id_col < len(row) else ""),
                "url": _safe_text(row[url_col] if url_col is not None and url_col < len(row) else ""),
            })

        return posts[:limit] if limit else posts
    finally:
        workbook.close()


def _blogger_name_from_path(root, path):
    root = Path(root).resolve()
    path = Path(path).resolve()
    relative = path.relative_to(root)
    if len(relative.parts) > 1:
        return relative.parts[0]

    return re.sub(r"(_\d{8})(_\d{6})?$", "", path.stem)


def _files_by_blogger(root):
    grouped = {}
    for path in _iter_excel_files(root):
        name = _blogger_name_from_path(root, path)
        grouped.setdefault(name, []).append(path)

    for paths in grouped.values():
        paths.sort(key=lambda p: p.stat().st_mtime, reverse=True)

    return grouped


def _latest_files_by_blogger(root):
    return {
        name: paths[0]
        for name, paths in _files_by_blogger(root).items()
        if paths
    }


def _dedupe_posts(posts):
    deduped = []
    seen = set()
    for post in posts:
        note_id = post.get("noteId") or ""
        title = post.get("title") or ""
        content = post.get("content") or ""
        key = f"id:{note_id}" if note_id else f"text:{title}\n{content}"
        if key in seen:
            continue
        seen.add(key)
        deduped.append(post)
    return deduped


def list_bloggers(root):
    bloggers = []
    for name, paths in sorted(_files_by_blogger(root).items()):
        all_posts = []
        latest_path = paths[0]
        for path in paths:
            all_posts.extend(_read_posts(path, limit=None))
        all_posts = _dedupe_posts(all_posts)

        bloggers.append({
            "name": name,
            "postCount": len(all_posts),
            "modifiedAt": latest_path.stat().st_mtime,
            "fileName": latest_path.name,
            "runDir": latest_path.parent.name,
            "fileCount": len(paths),
            "sampleTitles": [post["title"] for post in all_posts[:3] if post.get("title")],
        })
    return {"bloggers": bloggers}


def _find_blogger_files(root, blogger_name):
    """查找博主文件夹下的所有 Excel 文件"""
    root = Path(root).resolve()
    blogger_dir = root / blogger_name
    
    # 如果存在博主专属文件夹，读取该文件夹下所有 Excel
    if blogger_dir.exists() and blogger_dir.is_dir():
        excel_files = sorted(
            [f for f in blogger_dir.rglob("*.xlsx") if f.is_file() and not f.name.startswith("~$")],
            key=lambda p: p.stat().st_mtime,
            reverse=True  # 最新的在前
        )
        if excel_files:
            return excel_files
    
    # 回退到旧逻辑：在根目录查找，并兼容去掉日期后缀后的博主名
    grouped = _files_by_blogger(root)
    return grouped.get(blogger_name, [])


def read_blogger(root, name, limit):
    excel_files = _find_blogger_files(root, name)
    if not excel_files:
        raise FileNotFoundError(f"未找到博主 Excel: {name}")

    # 读取所有 Excel 文件的内容
    all_posts = []
    total_count = 0
    latest_mtime = 0
    
    for path in excel_files:
        posts = _read_posts(path, limit=None)
        all_posts.extend(posts)
        total_count += len(posts)
        mtime = path.stat().st_mtime
        if mtime > latest_mtime:
            latest_mtime = mtime

    all_posts = _dedupe_posts(all_posts)
    total_count = len(all_posts)

    # 如果有 limit，限制返回数量
    if limit and limit != 'all':
        all_posts = all_posts[:limit]

    return {
        "blogger": {
            "name": name,
            "postCount": total_count,
            "modifiedAt": latest_mtime,
            "fileCount": len(excel_files),
            "fileNames": [f.name for f in excel_files],
        },
        "posts": all_posts,
    }


def main():
    parser = argparse.ArgumentParser(description="Read crawler Excel files")
    parser.add_argument("--mode", choices=["list", "read"], required=True)
    parser.add_argument("--output-dir", default="./xiaohongshu_notes")
    parser.add_argument("--name", default="")
    parser.add_argument("--limit", default="30")
    args = parser.parse_args()

    try:
        if args.mode == "list":
            result = list_bloggers(args.output_dir)
        else:
            limit = None if args.limit == "all" else max(int(args.limit), 1)
            result = read_blogger(args.output_dir, args.name, limit)
        print(json.dumps(result, ensure_ascii=False))
    except Exception as exc:
        print(json.dumps({"error": str(exc)}, ensure_ascii=False))
        sys.exit(1)


if __name__ == "__main__":
    main()
